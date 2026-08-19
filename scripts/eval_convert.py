#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试题集 xlsx → 结构化 JSON 用例转换脚本。

用法:
  # KOL 测试题集（4 个 Sheet：一致性测试/立项判断/推广合作/设计反馈）
  python3 scripts/eval_convert.py "KOL数字孪生_测试题集_硬核测评KOL(2).xlsx" \
      --target kol --name "KOL数字孪生_测试题集" --out data/eval/test_cases_kol.json

  # 群体画像测试题集
  python3 scripts/eval_convert.py "AI模拟用户画像_测试题集_射击类用户(2).xlsx" \
      --target persona --name "群体画像测试题集" --out data/eval/test_cases_persona.json

  # 指定目标画像/KOL id（若题集里没有单独列）
  python3 scripts/eval_convert.py "....xlsx" --target kol --target-id 1 --out ...

说明:
  - 自动识别表头列名（题目/问题、维度/类别、参考答案/期望要点、画像/KOL/目标ID）。
  - 若真实 xlsx 表头与下面 COLUMN_ALIASES 不匹配，请按需增补别名，或手动整理成
    data/eval/ 下的标准 JSON 再跑 eval_run.py。
  - 生成的标准 JSON 结构见 data/eval/test_cases_sample.json。

依赖: openpyxl（pip3 install openpyxl）
"""
import argparse
import json
import os
import re
import sys

from openpyxl import load_workbook

# 表头别名（小写、去空格后匹配）。命中即作为该字段列。
COLUMN_ALIASES = {
    "question": ["题目", "问题", "提问", "question", "题"],
    "dimension": ["维度", "类别", "分类", "类型", "dimension", "category", "sheet"],
    "reference": ["参考答案", "期望要点", "参考答案要点", "参考", "答案", "要点", "reference", "answer"],
    "target_id": ["画像id", "画像ID", "画像", "kolid", "kol_id", "up主id", "目标id", "personaid", "persona_id", "id"],
}

# 表头行可能不在第 1 行，向上搜索这些关键词
DIMENSION_KEYWORDS = ["一致性", "立项", "推广", "设计", "反馈", "判断", "测评"]


def normalize(s: str) -> str:
    return re.sub(r"\s+", "", str(s or "")).lower()


def detect_columns(header_row: list) -> dict:
    """把表头行的每一列映射到 question/dimension/reference/target_id 四类之一。"""
    mapping = {}
    for idx, cell in enumerate(header_row):
        norm = normalize(cell)
        if not norm:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if field in mapping:
                continue
            if any(normalize(a) == norm or normalize(a) in norm for a in aliases):
                mapping[field] = idx
                break
    return mapping


def is_header_row(row: list) -> bool:
    """Check if row looks like a header by examining individual cells.
    Uses substring match per cell, but only if cells are short (header-like).
    """
    cells = [normalize(str(c)) for c in row if c]
    if len(cells) < 2:
        return False
    # Header cells are short; description cells are long
    short_cells = [c for c in cells if len(c) <= 10]
    if len(short_cells) < 2:
        return False
    # Check if any short cell contains a question or dimension alias
    q_aliases = [normalize(a) for a in COLUMN_ALIASES["question"]]
    d_aliases = [normalize(a) for a in COLUMN_ALIASES["dimension"]]
    has_question = any(
        any(a in c for a in q_aliases)
        for c in short_cells
    )
    has_dimension = any(
        any(a in c for a in d_aliases)
        for c in short_cells
    )
    return has_question or has_dimension

def sheet_to_cases(ws, target_id_default: int | None, seen_ids: set) -> list:
    """把单个 worksheet 转成 cases 列表。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 找表头行
    header_idx = 0
    for i, row in enumerate(rows[:5]):  # 前 5 行内找表头
        if is_header_row(list(row)):
            header_idx = i
            break

    cols = detect_columns(list(rows[header_idx]))
    if "question" not in cols:
        print(f"  ⚠ Sheet「{ws.title}」未识别到题目列，跳过", file=sys.stderr)
        return []

    cases = []
    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def cell(field):
            i = cols.get(field)
            if i is None or i >= len(row):
                return ""
            v = row[i]
            return str(v).strip() if v is not None else ""

        question = cell("question")
        if not question:
            continue

        tid = cell("target_id")
        tid = int(tid) if tid.lstrip("-").isdigit() else (target_id_default or None)

        case = {
            "id": None,  # 下面统一编号
            "dimension": cell("dimension") or ws.title,
            "target_id": tid,
            "question": question,
            "reference": cell("reference"),
        }
        cases.append(case)

    return cases


def main() -> int:
    parser = argparse.ArgumentParser(description="xlsx 测试题集 → JSON 用例")
    parser.add_argument("xlsx_path", help="测试题集 xlsx 路径")
    parser.add_argument("--target", choices=["persona", "kol"], default="persona", help="评测目标类型")
    parser.add_argument("--name", default=None, help="题集名（默认取文件名）")
    parser.add_argument("--target-id", type=int, default=None, help="当题集无目标列时，统一指定的 personaId/kolId")
    parser.add_argument("--sheet", default=None, help="只转换指定 Sheet 名（默认全部）")
    parser.add_argument("--out", required=True, help="输出 JSON 路径")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx_path, read_only=True, data_only=True)
    sheet_names = [args.sheet] if args.sheet else wb.sheetnames
    name = args.name or os.path.splitext(os.path.basename(args.xlsx_path))[0]

    cases: list = []
    seen = 0
    for sn in sheet_names:
        ws = wb[sn]
        sheet_cases = sheet_to_cases(ws, args.target_id, set())
        # 全题集统一编号：<维度缩写>-<序号>
        prefix = re.sub(r"[^A-Za-z0-9]", "", sn)[:3].upper() or "Q"
        for c in sheet_cases:
            seen += 1
            c["id"] = f"{prefix}-{seen:03d}"
        cases.extend(sheet_cases)
        print(f"Sheet「{sn}」: {len(sheet_cases)} 题")

    out = {
        "meta": {
            "name": name,
            "source_file": os.path.basename(args.xlsx_path),
            "target": args.target,
            "api_base": "http://localhost:3000",
            "dimensions": ["人设一致性", "专业准确性", "知识边界"],
        },
        "cases": cases,
    }

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"共 {len(cases)} 题 → {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
