#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按《数据提取规范 v1.3 §6.4 记录表（xlsx）提取规范》重新提取两份新手体验研究文件。

处理逻辑（§6.4 四类 Sheet 判定）：
  - ① 档案/样本（用户列表）        → respondents
  - ③ 访谈回答矩阵（列=受访者，行=问题）→ respondents + segments
  - ② 大纲 / ④ 行为观察·结论      → 不提取

输出：覆盖 data/群体画像v2.0/ 下对应 .json（替换旧的错误提取）。

用法：python3 scripts/extract_xlsx_record_tables.py
"""

import json
import os
import re

from openpyxl import load_workbook

BASE = "/Users/jessicajyan/Desktop/腾讯用户画像-data/data/虚拟用户-笔录 for 元培"
OUT = "/Users/jessicajyan/tengxun_yp_6Gang/data/群体画像v2.0"


# ---------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------
def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def clean_label(s):
    s = norm(s)
    s = re.sub(r"^[-\s]+", "", s)  # 去前导 '-'
    return s


# 结构性行（章节/引导语/任务/观察/阶段标题，无受访者数据），跳过
SKIP_EXACT = {
    "用户组别", "模块", "玩家编号", "玩家属性", "游玩顺序", "测试时间",
    "录像", "行为特征", "主要诉求",
    "测试前访谈", "测试后访谈", "游戏体验", "自由体验",
    "新手引导阶段自由体验", "新手引导阶段访谈",
    "第一次测试", "第二次测试", "第一次测试后访谈", "第二次测试后访谈", "特殊场景",
    "暖场", "暖场 & 信息甄别 & 初步了解", "暖场&信息甄别&初步了解",
    "第一局", "第二局", "第三局", "第四局", "第五局", "第六局", "第七局",
    "局前", "局中", "局后",
    "第一场", "第二场", "第三场",
    "局前行为记录", "局中异常行为记录", "局后行为记录", "继续游戏意愿度1-7",
    "新手目标", "伪装者", "有效性与目标感",
    "局内物资管理习惯（单排&组队）", "局内物资交换行为习惯和障碍（组队玩家）",
    "开局前配装行为习惯以及障碍点", "局外仓库系统中物资管理行为习惯以及障碍点",
}


def is_structure_row(a, b):
    for x in (a, b):
        x = norm(x)
        if not x:
            continue
        if x in SKIP_EXACT:
            return True
        if re.match(r"^[一二三四五六七八九十]+、", x):
            return True
        if x.startswith("【"):
            return True
        if re.match(r"^观察[:：]", x):
            return True
        if x.startswith("（") and re.search(r"暂停|如果|开始之前|进入结算", x):
            return True
    return False


def combine_question(a, b):
    a = clean_label(a)
    b = clean_label(b)
    if a and b and a != b:
        return f"{a}：{b}"
    return b or a


def new_respondent(speaker_id, source_file, display_name, group_code):
    return {
        "speaker_id": speaker_id,
        "source_file": source_file,
        "display_name": display_name,
        "group_code": group_code,
        "profile": {
            "name": display_name,
            "age": None,
            "gender": "",
            "occupation": "",
            "education": "",
        },
        "gaming_background": {
            "current_games": [],
            "platform": [],
            "experience_years": None,
            "genre_experience": [],
        },
        "background": {},
    }


# ---------------------------------------------------------------
# ③ 访谈回答矩阵：列头解析受访者 + 转置出 Segment
# ---------------------------------------------------------------
def parse_header_group_and_name(v):
    """从列头单元格解析 (group_code, display_name, 画像描述)。"""
    v = norm(v)
    # P1（G2、画像描述）  → group=G2, name=P1, desc=画像描述
    m = re.match(r"^(P\d+)\s*[（(]\s*(G\d+)\s*[、，,]?\s*(.*?)\s*[）)]?$", v)
    if m:
        return m.group(2), m.group(1), m.group(3).strip()
    # G1-P1 / G2 - P4 / G1-庄俊平 / G2余忠霖
    m = re.match(r"^(G\d+)\s*[-－]?\s*(.+)$", v)
    if m:
        return m.group(1), m.group(2), ""
    # P1 / P2
    m = re.match(r"^(P\d+)$", v)
    if m:
        return None, m.group(1), ""
    return None, v, ""


def build_matrix_group(source_file, ws, participant_cols, header_row=1,
                       attr_row=None, attr_key=None):
    """对一个 ③ 矩阵 Sheet 提取 respondents + segments。

    participant_cols: [(col_index, (group_code, display_name, 画像描述)), ...]
    header_row: 列头所在行（默认第 1 行）
    attr_row: 附加属性行（如「暖场/玩家属性」行），写入 background
    """
    respondents = []
    segments = []

    # 1) 构建受访者
    for ci, (g, dname, desc) in participant_cols:
        sid = f"P{len(respondents) + 1:03d}"
        r = new_respondent(sid, source_file, dname, g or "")
        # 列头里带的画像描述（如「只玩和平钢枪男…」）
        if desc:
            r["background"]["画像"] = desc
        # 附加属性（玩家类型标签 / 游玩顺序等）
        if attr_row is not None:
            av = ws.cell(attr_row, ci).value
            if av is not None and norm(av):
                r["background"][attr_key or "属性"] = norm(av)
        respondents.append(r)

    # 2) 转置出 Segment
    max_row = ws.max_row
    for ri in range(1, max_row + 1):
        a = ws.cell(ri, 1).value
        b = ws.cell(ri, 2).value
        if is_structure_row(a, b):
            continue
        q = combine_question(a, b)
        if not q:
            continue
        for ci, (g, dname, desc) in participant_cols:
            cell = ws.cell(ri, ci).value
            text = norm(cell)
            if not text or len(text) < 2:
                continue
            # 对应第几个受访者
            sid = respondents[participant_cols.index((ci, (g, dname, desc)))]["speaker_id"]
            segments.append({
                "speaker_id": sid,
                "speaker_role": "interviewee",
                "preceding_question": q,
                "original_text": text,
            })

    return respondents, segments


# ---------------------------------------------------------------
# 文件一：用户记录表.xlsx（枪战类长线新手体验研究）
# ---------------------------------------------------------------
def parse_gunbattle():
    rel = "枪战类长线新手体验研究/用户记录表.xlsx"
    path = os.path.join(BASE, rel)
    wb = load_workbook(path, data_only=True)
    all_respondents = []
    all_segments = []

    # --- 研究一（塔科夫like，美国玩家）：用户列表 → respondents ---
    src1 = f"{rel}/用户列表"
    ws = wb["用户列表"]
    header_row = None
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c) if c is not None else "" for c in r])
    for i, row in enumerate(rows):
        if any("ID" == norm(c) or "Name" in str(c) for c in row):
            header_row = i
            break
    if header_row is not None:
        header = rows[header_row]
        # 列索引（1-based）
        col_id = 0
        col_name = 1
        col_seg = 3
        col_gender = 5
        col_age = 6
        col_eth = 7
        col_plat = 8
        # 游戏清单列（x 标记）
        game_cols = [c for c in range(10, 31)]
        for row in rows[header_row + 1:]:
            rid = norm(row[col_id]) if len(row) > col_id else ""
            name = norm(row[col_name]) if len(row) > col_name else ""
            if not rid:
                continue
            gm = re.match(r"^(G\d+)", rid)
            group = gm.group(1) if gm else ""
            gender = norm(row[col_gender]) if len(row) > col_gender else ""
            age_raw = norm(row[col_age]) if len(row) > col_age else ""
            try:
                age = int(age_raw) if age_raw else None
            except ValueError:
                age = None
            eth = norm(row[col_eth]) if len(row) > col_eth else ""
            plat = norm(row[col_plat]) if len(row) > col_plat else ""
            games = []
            for c in game_cols:
                if len(row) > c and norm(row[c]).lower() == "x":
                    gh = norm(header[c]) if len(header) > c else ""
                    if gh:
                        games.append(gh)

            sid = f"P{len(all_respondents) + 1:03d}"
            r = new_respondent(sid, src1, name, group)
            r["profile"].update({
                "name": name, "age": age, "gender": gender,
                "ethnicity": eth,
            })
            r["gaming_background"]["platform"] = [p.strip() for p in plat.split(",") if p.strip()]
            r["gaming_background"]["current_games"] = games
            r["background"]["roster_id"] = rid
            all_respondents.append(r)

    # --- 研究二（ABI vs DF，中国玩家）：ABIvsDF记录表 → respondents + segments ---
    src2 = f"{rel}/ABIvsDF记录表"
    ws = wb["ABIvsDF记录表"]
    # 列头 r1 = P1..P8，属性 r2 = G1 CS 卢 等
    participant_cols = []
    for ci in range(3, 11):
        h = norm(ws.cell(1, ci).value)
        if not h:
            continue
        g, dname, desc = parse_header_group_and_name(ws.cell(2, ci).value)
        if not dname:
            dname = h
        participant_cols.append((ci, (g, dname, desc)))
    # 修正：列头 r1 是 P1..P8，属性在 r2；用 r2 的 group 解析结果，若失败则用 r1
    # （parse_header_group_and_name 已传 r2 单元格，返回 (group, name, desc)）
    r2_resp, r2_seg = build_matrix_group(
        src2, ws, participant_cols, header_row=1, attr_row=3, attr_key="游玩顺序"
    )
    # 重新编号（接在研究一之后）
    offset = len(all_respondents)
    remap = {}
    for i, r in enumerate(r2_resp):
        new_sid = f"P{offset + i + 1:03d}"
        remap[r["speaker_id"]] = new_sid
        r["speaker_id"] = new_sid
        all_respondents.append(r)
    for s in r2_seg:
        s["speaker_id"] = remap[s["speaker_id"]]
        all_segments.append(s)

    wb.close()
    return rel, all_respondents, all_segments


# ---------------------------------------------------------------
# 文件二：整合记录表.xlsx（生存撤离类新手引导体验研究）
# ---------------------------------------------------------------
def parse_shengcun():
    rel = "生存撤离类新手引导体验研究/整合记录表.xlsx"
    path = os.path.join(BASE, rel)
    wb = load_workbook(path, data_only=True)
    all_respondents = []
    all_segments = []

    def add_group(sheet_name, cols, attr_row=None, attr_key=None):
        nonlocal all_respondents, all_segments
        src = f"{rel}/{sheet_name}"
        ws = wb[sheet_name]
        offset = len(all_respondents)
        resp, seg = build_matrix_group(src, ws, cols, header_row=1,
                                       attr_row=attr_row, attr_key=attr_key)
        remap = {}
        for i, r in enumerate(resp):
            new_sid = f"P{offset + i + 1:03d}"
            remap[r["speaker_id"]] = new_sid
            r["speaker_id"] = new_sid
            all_respondents.append(r)
        for s in seg:
            s["speaker_id"] = remap[s["speaker_id"]]
            all_segments.append(s)

    # 一期暗区：列 5-10 = G2-P4/G2-P3/G1-P2/G3-P5/G3-P6/G1-P1；r2=玩家类型标签
    ws = wb["【一期】暗区"]
    cols = []
    for ci in range(5, 11):
        h = norm(ws.cell(1, ci).value)
        if h:
            g, dname, desc = parse_header_group_and_name(h)
            cols.append((ci, (g, dname or h, desc)))
    add_group("【一期】暗区", cols, attr_row=2, attr_key="类型")

    # 二期塔科夫：列 3-6 = G1-庄俊平/G2余忠霖/G2-包叶喆/G3-张凤武（无附加属性行）
    ws = wb["【二期】塔科夫"]
    cols = []
    for ci in range(3, 7):
        h = norm(ws.cell(1, ci).value)
        if h:
            g, dname, desc = parse_header_group_and_name(h)
            cols.append((ci, (g, dname or h, desc)))
    add_group("【二期】塔科夫", cols)

    # 三期新手综合：列 3-8 = P1（G2、画像描述）…P6（G1、…）；画像在列头括号内
    ws = wb["【三期】新手综合"]
    cols = []
    for ci in range(3, 9):
        h = norm(ws.cell(1, ci).value)
        if h:
            g, dname, desc = parse_header_group_and_name(h)
            cols.append((ci, (g, dname or h, desc)))
    add_group("【三期】新手综合", cols)

    wb.close()
    return rel, all_respondents, all_segments


# ---------------------------------------------------------------
# 输出
# ---------------------------------------------------------------
def write_output(rel, respondents, segments):
    out_rel = os.path.splitext(rel)[0] + ".json"
    out_path = os.path.join(OUT, out_rel)
    output = {
        "meta": {
            "version": "v2.2",
            "processing_date": "2026-08-26",
            "source_file": rel,
            "source_type": "记录表（xlsx）",
            "participant_count": len(respondents),
            "segment_count": len(segments),
            "processing_notes": [
                "§6.4 记录表（xlsx）四类 Sheet 判定：①档案 ②大纲 ③访谈回答矩阵 ④观察/结论",
                "③ 访谈回答矩阵转置：列头→speaker_id，行头→preceding_question，非空单元格→segment",
                "④ 行为观察/结论类不产 Segment",
                "同一 workbook 不同批人拆分为独立 source_file（sheet 后缀）",
            ],
        },
        "respondents": respondents,
        "segments": [
            {
                "segment_id": j + 1,
                "source_file": s["speaker_id"],  # 占位，下面修正
                "segment_index": j + 1,
                "speaker_id": s["speaker_id"],
                "speaker_role": s.get("speaker_role", "interviewee"),
                "preceding_question": s["preceding_question"],
                "original_text": s["original_text"],
                "cleaned_text": None,
                "char_count": len(s["original_text"]),
            }
            for j, s in enumerate(segments)
        ],
    }
    # 修正 source_file：从 respondent 反查
    sid_to_source = {r["speaker_id"]: r["source_file"] for r in respondents}
    for seg in output["segments"]:
        seg["source_file"] = sid_to_source.get(seg["speaker_id"], rel)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    return out_path, len(respondents), len(segments)


def main():
    for parser in (parse_gunbattle, parse_shengcun):
        rel, resp, seg = parser()
        out_path, nr, ns = write_output(rel, resp, seg)
        print(f"OK {rel} -> {out_path}")
        print(f"   respondents={nr}, segments={ns}")


if __name__ == "__main__":
    main()
